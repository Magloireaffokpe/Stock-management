"""
Tests — app reports
Couvre : Dashboard KPIs, graphiques (daily/monthly/catégories),
         top produits, valeur stock, exports Excel
"""

from datetime import date, timedelta
from decimal import Decimal
import io

from django.utils import timezone
from rest_framework import status

from apps.sales.models import Sale, SaleItem
from .base import BaseTestCase
from .factories import make_product, make_client, make_sale, make_category, make_store

# ═══════════════════════════════════════════════════════════════════
#  TESTS INTÉGRATION — Dashboard KPIs
# ═══════════════════════════════════════════════════════════════════


class DashboardKPITest(BaseTestCase):

    def test_accessible_aux_utilisateurs_authentifiés(self):
        r_emp = self.employee_client.get("/api/reports/dashboard/")
        self.assertEqual(r_emp.status_code, status.HTTP_200_OK)

        r_admin = self.admin_client.get("/api/reports/dashboard/")
        self.assertEqual(r_admin.status_code, status.HTTP_200_OK)

    def test_non_accessible_sans_auth(self):
        r = self.client.get("/api/reports/dashboard/")
        self.assertEqual(r.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_structure_réponse_complète(self):
        r = self.admin_client.get("/api/reports/dashboard/")
        self.assertEqual(r.status_code, status.HTTP_200_OK)

        # Vérifier les blocs principaux
        for bloc in ["today", "week", "month", "year", "stock"]:
            self.assertIn(bloc, r.data, f"Bloc manquant : {bloc}")

        # Vérifier les champs dans 'today'
        for champ in [
            "revenue",
            "count",
            "variation_revenue",
            "variation_count",
        ]:
            self.assertIn(champ, r.data["today"])

        # Vérifier les champs dans 'stock'
        for champ in [
            "total_products",
            "out_of_stock",
            "critical",
            "low",
            "unread_alerts",
        ]:
            self.assertIn(champ, r.data["stock"])

    def test_kpi_today_reflète_vente_du_jour(self):
        """Une vente créée aujourd'hui doit apparaître dans today.revenue"""
        p = make_product(selling_price=75000, stock_quantity=10)
        make_sale(
            self.admin, items=[{"product": p, "quantity": 1, "unit_price": 75000}]
        )

        r = self.admin_client.get("/api/reports/dashboard/")
        self.assertGreaterEqual(r.data["today"]["revenue"], 75000)
        self.assertGreaterEqual(r.data["today"]["count"], 1)

    def test_kpi_exclut_ventes_annulées(self):
        """Les ventes annulées ne doivent pas être comptabilisées dans le CA"""
        p = make_product(selling_price=100000, stock_quantity=10)
        sale = make_sale(
            self.admin, items=[{"product": p, "quantity": 1, "unit_price": 100000}]
        )

        r_avant = self.admin_client.get("/api/reports/dashboard/")
        revenue_avant = r_avant.data["today"]["revenue"]

        # Annuler la vente
        self.admin_client.post(f"/api/sales/sales/{sale.pk}/cancel/")

        r_après = self.admin_client.get("/api/reports/dashboard/")
        revenue_après = r_après.data["today"]["revenue"]

        self.assertLess(revenue_après, revenue_avant)

    def test_stock_counts_corrects(self):
        """Le dashboard doit compter correctement les produits en rupture/faible/critique"""
        make_product(name="KPI OK 1", stock_quantity=20)
        make_product(name="KPI Low", stock_quantity=3, low_stock_threshold=5)
        make_product(name="KPI Critical", stock_quantity=2)  # seuil critical=2
        make_product(name="KPI Out", stock_quantity=0)

        r = self.admin_client.get("/api/reports/dashboard/")
        stock_data = r.data["stock"]

        self.assertGreaterEqual(stock_data["out_of_stock"], 1)
        self.assertGreaterEqual(stock_data["critical"], 1)
        self.assertGreaterEqual(stock_data["low"], 1)
        self.assertGreaterEqual(stock_data["total_products"], 4)


# ═══════════════════════════════════════════════════════════════════
#  TESTS INTÉGRATION — Graphiques
# ═══════════════════════════════════════════════════════════════════


class DailySalesChartTest(BaseTestCase):

    def test_accessible_aux_utilisateurs_authentifiés(self):
        r_emp = self.employee_client.get("/api/reports/charts/daily/")
        self.assertEqual(r_emp.status_code, status.HTTP_200_OK)

        r_admin = self.admin_client.get("/api/reports/charts/daily/")
        self.assertEqual(r_admin.status_code, status.HTTP_200_OK)
        self.assertIsInstance(r_admin.data, list)

    def test_retourne_7_jours_par_défaut(self):
        r = self.admin_client.get("/api/reports/charts/daily/")
        self.assertEqual(len(r.data), 7)

    def test_paramètre_days_personnalisé(self):
        r = self.admin_client.get("/api/reports/charts/daily/?days=30")
        self.assertEqual(len(r.data), 30)

    def test_structure_chaque_entrée(self):
        r = self.admin_client.get("/api/reports/charts/daily/?days=3")
        for entry in r.data:
            self.assertIn("date", entry)
            self.assertIn("revenue", entry)
            self.assertIn("count", entry)

    def test_jours_sans_ventes_remplis_avec_zéros(self):
        """Les jours sans ventes doivent avoir revenue=0 et count=0"""
        r = self.admin_client.get("/api/reports/charts/daily/?days=7")
        yesterday = (date.today() - timedelta(days=2)).isoformat()
        yesterday_data = next((d for d in r.data if d["date"] == yesterday), None)
        if yesterday_data:
            self.assertEqual(yesterday_data["count"], 0)
            self.assertEqual(yesterday_data["revenue"], 0)

    def test_today_flagged(self):
        r = self.admin_client.get("/api/reports/charts/daily/?days=7")
        today_str = date.today().isoformat()
        today_entry = next((d for d in r.data if d["date"] == today_str), None)
        if today_entry:
            self.assertTrue(today_entry.get("is_today"))


class MonthlySalesChartTest(BaseTestCase):

    def test_accessible_aux_utilisateurs_authentifiés(self):
        r_emp = self.employee_client.get("/api/reports/charts/monthly/")
        self.assertEqual(r_emp.status_code, status.HTTP_200_OK)

        r_admin = self.admin_client.get("/api/reports/charts/monthly/")
        self.assertEqual(r_admin.status_code, status.HTTP_200_OK)

    def test_structure_entrée(self):
        r = self.admin_client.get("/api/reports/charts/monthly/")
        for entry in r.data:
            self.assertIn("month", entry)
            self.assertIn("label", entry)
            self.assertIn("revenue", entry)
            self.assertIn("count", entry)


class CategorySalesChartTest(BaseTestCase):

    def test_accessible_aux_utilisateurs_authentifiés(self):
        r_emp = self.employee_client.get("/api/reports/charts/categories/")
        self.assertEqual(r_emp.status_code, status.HTTP_200_OK)

        r_admin = self.admin_client.get("/api/reports/charts/categories/")
        self.assertEqual(r_admin.status_code, status.HTTP_200_OK)

    def test_retourne_catégories_avec_ventes(self):
        cat = make_category(name="Cat Chart Test")
        p = make_product(
            name="Prod Chart", category=cat, stock_quantity=10, selling_price=20000
        )
        make_sale(
            self.admin, items=[{"product": p, "quantity": 1, "unit_price": 20000}]
        )

        r = self.admin_client.get("/api/reports/charts/categories/")
        noms = [e.get("product__category__name") for e in r.data]
        self.assertIn("Cat Chart Test", noms)


class TopProductsTest(BaseTestCase):

    def test_accessible_aux_utilisateurs_authentifiés(self):
        r_emp = self.employee_client.get("/api/reports/top-products/")
        self.assertEqual(r_emp.status_code, status.HTTP_200_OK)

        r_admin = self.admin_client.get("/api/reports/top-products/")
        self.assertEqual(r_admin.status_code, status.HTTP_200_OK)

    def test_top_produit_trié_par_ca(self):
        p_gros = make_product(
            name="Top Vendeur", selling_price=100000, stock_quantity=20
        )
        p_petit = make_product(
            name="Petit Vendeur", selling_price=5000, stock_quantity=20
        )
        make_sale(
            self.admin, items=[{"product": p_gros, "quantity": 3, "unit_price": 100000}]
        )
        make_sale(
            self.admin, items=[{"product": p_petit, "quantity": 1, "unit_price": 5000}]
        )

        r = self.admin_client.get("/api/reports/top-products/?limit=10")
        noms = [e["product__name"] for e in r.data]
        idx_gros = noms.index("Top Vendeur") if "Top Vendeur" in noms else -1
        idx_petit = noms.index("Petit Vendeur") if "Petit Vendeur" in noms else 9999
        self.assertLess(idx_gros, idx_petit)

    def test_limite_nombre_résultats(self):
        for i in range(12):
            p = make_product(name=f"Top P {i}", stock_quantity=10)
            make_sale(
                self.admin, items=[{"product": p, "quantity": 1, "unit_price": 5000}]
            )
        r = self.admin_client.get("/api/reports/top-products/?limit=5")
        self.assertLessEqual(len(r.data), 5)


class RecentSalesTest(BaseTestCase):

    def test_accessible_aux_utilisateurs_authentifiés(self):
        r_emp = self.employee_client.get("/api/reports/dashboard/recent-sales/")
        self.assertEqual(r_emp.status_code, status.HTTP_200_OK)

        r_admin = self.admin_client.get("/api/reports/dashboard/recent-sales/")
        self.assertEqual(r_admin.status_code, status.HTTP_200_OK)

    def test_retourne_max_5_ventes(self):
        p = make_product(stock_quantity=100)
        for i in range(7):
            make_sale(
                self.admin, items=[{"product": p, "quantity": 1, "unit_price": 5000}]
            )
        r = self.admin_client.get("/api/reports/dashboard/recent-sales/")
        self.assertLessEqual(len(r.data), 5)

    def test_ventes_triées_plus_récentes_en_premier(self):
        p = make_product(stock_quantity=50)
        s1 = make_sale(
            self.admin, items=[{"product": p, "quantity": 1, "unit_price": 10000}]
        )
        s2 = make_sale(
            self.admin, items=[{"product": p, "quantity": 1, "unit_price": 10000}]
        )
        r = self.admin_client.get("/api/reports/dashboard/recent-sales/")
        ids = [s["id"] for s in r.data]
        if s1.pk in ids and s2.pk in ids:
            self.assertLess(ids.index(s2.pk), ids.index(s1.pk))

    def test_historique_dashboard_employe_limite_a_ses_7_derniers_jours(self):
        product = make_product(stock_quantity=20)
        own_recent = make_sale(
            self.employee, items=[{"product": product, "quantity": 1, "unit_price": 10000}]
        )
        other_recent = make_sale(
            self.admin, items=[{"product": product, "quantity": 1, "unit_price": 10000}]
        )
        own_old = make_sale(
            self.employee, items=[{"product": product, "quantity": 1, "unit_price": 10000}]
        )
        Sale.objects.filter(pk=own_old.pk).update(
            sale_date=timezone.now() - timedelta(days=7)
        )

        response = self.employee_client.get("/api/reports/dashboard/recent-sales/")
        ids = [sale["id"] for sale in response.data]

        self.assertIn(own_recent.pk, ids)
        self.assertNotIn(other_recent.pk, ids)
        self.assertNotIn(own_old.pk, ids)


# ═══════════════════════════════════════════════════════════════════
#  TESTS INTÉGRATION — Exports Excel
# ═══════════════════════════════════════════════════════════════════


class ExcelExportTest(BaseTestCase):

    def test_export_ventes_restreint_aux_admins(self):
        p = make_product(stock_quantity=10)
        make_sale(
            self.admin, items=[{"product": p, "quantity": 1, "unit_price": 10000}]
        )

        r_emp = self.employee_client.get("/api/reports/export/sales/")
        self.assertEqual(r_emp.status_code, status.HTTP_403_FORBIDDEN)

        r_admin = self.admin_client.get("/api/reports/export/sales/")
        self.assertEqual(r_admin.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", r_admin.get("Content-Type", ""))

    def test_export_ventes_retourne_xlsx(self):
        p = make_product(stock_quantity=10)
        make_sale(
            self.admin, items=[{"product": p, "quantity": 1, "unit_price": 10000}]
        )
        r = self.admin_client.get("/api/reports/export/sales/")
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        content_type = r.get("Content-Type", "")
        self.assertIn("spreadsheetml", content_type)
        # Vérifier que le fichier contient des données (pas vide)
        self.assertGreater(len(r.content), 1000)

    def test_export_produits_retourne_xlsx(self):
        make_product(name="Export Test Product", stock_quantity=5)
        r = self.admin_client.get("/api/reports/export/products/")
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", r.get("Content-Type", ""))

    def test_export_mouvements_retourne_xlsx(self):
        from apps.stock.models import StockMovement

        p = make_product(stock_quantity=10)
        StockMovement.objects.create(
            product=p,
            movement_type="initial",
            quantity=10,
            stock_before=0,
            stock_after=10,
        )
        r = self.admin_client.get("/api/reports/export/movements/")
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", r.get("Content-Type", ""))

    def test_export_ventes_avec_filtre_date(self):
        today = date.today().isoformat()
        r = self.admin_client.get(
            f"/api/reports/export/sales/?date_from={today}&date_to={today}"
        )
        self.assertEqual(r.status_code, status.HTTP_200_OK)

    def test_export_ventes_contient_les_produits_vendus(self):
        from openpyxl import load_workbook

        p = make_product(name="Produit Export Visible", stock_quantity=10)
        make_sale(
            self.admin,
            items=[{"product": p, "quantity": 3, "unit_price": 5000}],
        )
        r = self.admin_client.get("/api/reports/export/sales/")
        self.assertEqual(r.status_code, status.HTTP_200_OK)

        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual(len(wb.sheetnames), 1)
        ws = wb.active
        products = [row[3].value for row in ws.iter_rows(min_row=2)]
        self.assertIn("Produit Export Visible", products)
        quantities = [row[4].value for row in ws.iter_rows(min_row=2)]
        self.assertIn(3, quantities)

    def test_export_non_accessible_sans_auth(self):
        r = self.client.get("/api/reports/export/sales/")
        self.assertEqual(r.status_code, status.HTTP_401_UNAUTHORIZED)


class StoreFilteredReportsTest(BaseTestCase):
    def setUp(self):
        super().setUp()
        self.store1 = make_store(name='Boutique Alpha')
        self.store2 = make_store(name='Boutique Beta')

        self.cat1 = make_category(store=self.store1, name="Boutique 1 Root")
        self.cat2 = make_category(store=self.store2, name="Boutique 2 Root")

        self.p1 = make_product(category=self.cat1, store=self.store1, selling_price=10000, stock_quantity=10)
        self.p2 = make_product(category=self.cat2, store=self.store2, selling_price=20000, stock_quantity=5)

        # Create a sale for store 1
        make_sale(
            self.admin,
            items=[{"product": self.p1, "quantity": 1, "unit_price": 10000}],
            payment_method="cash"
        )
        # Create a sale for store 2
        make_sale(
            self.admin,
            items=[{"product": self.p2, "quantity": 1, "unit_price": 20000}],
            payment_method="mtn"
        )

    def test_dashboard_kpi_with_store_filter(self):
        # Without filter (Global)
        r_global = self.admin_client.get("/api/reports/dashboard/")
        self.assertEqual(r_global.data["today"]["revenue"], 30000)
        self.assertEqual(r_global.data["stock"]["total_products"], 2)

        # Filtered by store 1
        r_store1 = self.admin_client.get(f"/api/reports/dashboard/?store={self.store1.pk}")
        self.assertEqual(r_store1.data["today"]["revenue"], 10000)
        self.assertEqual(r_store1.data["stock"]["total_products"], 1)

    def test_daily_sales_chart_with_store_filter(self):
        today_str = date.today().isoformat()

        # Filtered by store 2
        r = self.admin_client.get(f"/api/reports/charts/daily/?store={self.store2.pk}&days=1")
        entry = next(d for d in r.data if d["date"] == today_str)
        self.assertEqual(entry["revenue"], 20000)

    def test_monthly_sales_chart_with_store_filter(self):
        r = self.admin_client.get(f"/api/reports/charts/monthly/?store={self.store1.pk}")
        self.assertEqual(r.data[0]["revenue"], 10000)

    def test_top_products_with_store_filter(self):
        r = self.admin_client.get(f"/api/reports/top-products/?store={self.store2.pk}")
        self.assertEqual(len(r.data), 1)
        self.assertEqual(r.data[0]["product__name"], self.p2.name)

    def test_category_sales_with_store_filter(self):
        r = self.admin_client.get(f"/api/reports/charts/categories/?store={self.store1.pk}")
        self.assertEqual(len(r.data), 1)
        self.assertEqual(r.data[0]["product__category__name"], self.cat1.name)

    def test_recent_sales_with_store_filter(self):
        r = self.admin_client.get(f"/api/reports/dashboard/recent-sales/?store={self.store1.pk}")
        self.assertEqual(len(r.data), 1)
        # SaleListSerializer is lightweight — verify by total_amount
        self.assertEqual(int(r.data[0]["total_amount"]), 10000)

    def test_excel_export_products_with_store_filter(self):
        # Store 1
        r1 = self.admin_client.get(f"/api/reports/export/products/?store={self.store1.pk}")
        self.assertEqual(r1.status_code, status.HTTP_200_OK)

        # Global
        r2 = self.admin_client.get("/api/reports/export/products/")
        self.assertEqual(r2.status_code, status.HTTP_200_OK)

    def test_excel_export_sales_with_store_filter(self):
        r1 = self.admin_client.get(f"/api/reports/export/sales/?store={self.store1.pk}")
        self.assertEqual(r1.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", r1.get("Content-Type", ""))
        # Le nom de fichier doit contenir la boutique
        self.assertIn("boutique-alpha", r1["Content-Disposition"])

    def test_excel_export_movements_with_store_filter(self):
        from apps.stock.models import StockMovement
        for store, prod in [(self.store1, self.p1), (self.store2, self.p2)]:
            StockMovement.objects.create(
                product=prod,
                movement_type="initial",
                quantity=5,
                stock_before=0,
                stock_after=5,
            )
        r1 = self.admin_client.get(f"/api/reports/export/movements/?store={self.store2.pk}")
        self.assertEqual(r1.status_code, status.HTTP_200_OK)
        self.assertIn("spreadsheetml", r1.get("Content-Type", ""))
        self.assertIn("boutique-beta", r1["Content-Disposition"])

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r1.content))
        ws = wb.active
        names = [row[1].value for row in ws.iter_rows(min_row=2)]
        self.assertIn(self.p2.name, names)
        self.assertNotIn(self.p1.name, names)
