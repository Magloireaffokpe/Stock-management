"""
Exporteurs de données vers Excel (openpyxl).
Appelés depuis reports/views.py pour les exports.
"""
import io
from datetime import date

from django.http import HttpResponse
from django.utils.text import slugify

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


HEADER_FILL  = PatternFill('solid', fgColor='1A2B4A') if OPENPYXL_AVAILABLE else None
HEADER_FONT  = Font(color='FFFFFF', bold=True) if OPENPYXL_AVAILABLE else None
ACCENT_FILL  = PatternFill('solid', fgColor='EFF6FF') if OPENPYXL_AVAILABLE else None


def _build_filename(base, store_label=None):
    """Nom de fichier distinct selon l'export et la boutique choisie.
    Ex : ventes_boutique-a_2026-08-16.xlsx / ventes_2026-08-16.xlsx"""
    today = date.today().isoformat()
    if store_label:
        label = slugify(store_label) or 'boutique'
        return f'{base}_{label}_{today}.xlsx'
    return f'{base}_{today}.xlsx'


def _auto_width(ws):
    """Ajuste automatiquement la largeur des colonnes"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def export_sales_excel(queryset, store_label=None):
    """Export de l'historique des ventes avec le détail des produits vendus.

    Une seule feuille : une ligne par article vendu. Les colonnes propres
    à la vente (facture, date, client, total, paiement, statut, caissier)
    sont fusionnées verticalement sur les lignes d'articles de cette vente.
    """
    if not OPENPYXL_AVAILABLE:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Historique ventes'

    headers = ['N° Facture', 'Date', 'Client', 'Produit', 'Qté',
               'Prix unitaire', 'Total ligne', 'Total', 'Statut', 'Caissier']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    # Colonnes propres à la vente (1-based) fusionnées sur les lignes d'articles
    SALE_MERGE_COLS = [1, 2, 3, 8, 9, 10]
    row = 2

    for sale in queryset:
        items = list(sale.items.all())
        if not items:
            continue

        start = row
        for item in items:
            ws.append([
                sale.invoice_number,
                sale.sale_date.strftime('%d/%m/%Y %H:%M'),
                sale.client.full_name if sale.client else 'Client comptoir',
                item.product_name,
                item.quantity,
                float(item.unit_price),
                float(item.subtotal),
                float(sale.total_amount),
                'Annulée' if sale.is_cancelled else 'Validée',
                sale.created_by.get_full_name() if sale.created_by else '',
            ])
            row += 1

        if row - 1 > start:
            for col in SALE_MERGE_COLS:
                ws.merge_cells(start_row=start, start_column=col,
                               end_row=row - 1, end_column=col)

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = _build_filename('ventes', store_label)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_products_excel(queryset, store_id=None, store_label=None):
    """Export du catalogue produits"""
    if not OPENPYXL_AVAILABLE:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Catalogue produits'

    def get_category_path(category):
        if not category:
            return ""
        path_parts = []
        current = category
        while current:
            path_parts.insert(0, current.name)
            current = current.parent
        return " > ".join(path_parts)

    if store_id:
        headers = ['SKU', 'Nom', 'Chemin catégorie', 'Fournisseur', 'État',
                   'Prix vente', 'Stock', 'Statut stock', 'Actif']
    else:
        headers = ['Boutique', 'SKU', 'Nom', 'Chemin catégorie', 'Fournisseur', 'État',
                   'Prix vente', 'Stock', 'Statut stock', 'Actif']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')

    for product in queryset:
        cat_path = get_category_path(product.category)
        selling_price = float(product.selling_price) if product.selling_price is not None else 0.0

        row_data = []
        if not store_id:
            row_data.append(product.store.name if product.store else '')

        row_data.extend([
            product.sku,
            product.name,
            cat_path,
            product.supplier.name if product.supplier else '',
            product.get_condition_display(),
            selling_price,
            product.stock_quantity,
            product.stock_status,
            'Oui' if product.is_active else 'Non',
        ])
        ws.append(row_data)

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = _build_filename('produits', store_label)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_stock_movements_excel(queryset, store_label=None):
    """Export des mouvements de stock"""
    """Export des mouvements de stock"""
    if not OPENPYXL_AVAILABLE:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mouvements stock'

    headers = ['Date', 'Produit', 'SKU', 'Type mouvement', 'Qté',
               'Stock avant', 'Stock après', 'Référence', 'Note', 'Utilisateur']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for m in queryset:
        ws.append([
            m.created_at.strftime('%d/%m/%Y %H:%M'),
            m.product.name,
            m.product.sku,
            m.get_movement_type_display(),
            m.quantity,
            m.stock_before,
            m.stock_after,
            m.reference,
            m.note,
            m.created_by.get_full_name() if m.created_by else 'Système',
        ])

    _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = _build_filename('mouvements_stock', store_label)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
